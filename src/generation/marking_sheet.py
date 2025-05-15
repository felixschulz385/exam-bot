import pandas as pd
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)

class MarkingSheetGenerator:
    """Generates marking sheets for easy grading of exams."""
    
    def __init__(self, output_dir: Path = Path("output")):
        """Initialize marking sheet generator.
        
        Args:
            output_dir: Directory to save marking sheets
        """
        self.output_dir = Path(output_dir)
        
        # Create output directory if it doesn't exist
        if not self.output_dir.exists():
            self.output_dir.mkdir(parents=True)
            logger.info(f"Created output directory: {self.output_dir}")
    
    def generate_marking_sheet(self, 
                          questions: pd.DataFrame, 
                          config: Dict[str, Any],
                          timestamp: str = None,
                          answer_mappings: Optional[Dict] = None) -> Path:
        """Generate a marking sheet in Excel format with support for randomized answers.
    
        Args:
            questions: DataFrame of selected questions
            config: Configuration dictionary
            timestamp: Optional timestamp for filename
            answer_mappings: Mappings of original to randomized answers
    
        Returns:
            Path to the generated marking sheet
        """
        # Check if answers were randomized
        randomize_answers = config.get("document", {}).get("randomize_answers", False)
        
        # Create a new DataFrame for the marking sheet
        marking_data = []
    
        # Get type points for adding to questions
        grading_config = config.get("grading", {})
        type_points = grading_config.get("type_points", {})
    
        # Process questions in the order they appear in the exam
        for question_num, (idx, row) in enumerate(questions.iterrows(), 1):
            # Extract question information
            question_text = row['Question']
            question_type = row['Type']
            topic = row['Topic']
            
            # Get correct answer number
            correct_answer_num = row.get('Correct', None)
            
            # Get correct answer letter - apply randomization if enabled
            correct_answer_letter = ""
            if correct_answer_num is not None:
                if randomize_answers and answer_mappings and question_num in answer_mappings:
                    # Use the randomized letter from the mapping
                    correct_answer_letter = answer_mappings[question_num].get(
                        correct_answer_num, 
                        chr(64 + correct_answer_num)  # Fallback to original if not found
                    )
                else:
                    # Use original letter (A, B, C, D)
                    correct_answer_letter = chr(64 + correct_answer_num)
        
            # Get correct answer text
            correct_answer_text = ""
            if correct_answer_num is not None and 1 <= correct_answer_num <= 4:
                answer_key = f'Answer {correct_answer_num}'
                if answer_key in row:
                    correct_answer_text = row[answer_key]
                
            # Calculate points for this question
            points = type_points.get(question_type, 1)
            
            # Format question text (truncate if too long)
            max_question_length = 50
            if len(question_text) > max_question_length:
                question_text = question_text[:max_question_length] + "..."
            
            # Add to marking data
            marking_data.append({
                'Question #': question_num,
                'Question Type': question_type,
                'Topic': topic,
                'Correct Answer #': correct_answer_num,
                'Correct Answer Letter': correct_answer_letter,
                'Correct Answer Text': correct_answer_text,
                'Points': points,
                'Question Text': question_text
            })
    
        # Create DataFrame
        marking_df = pd.DataFrame(marking_data)
    
        # Calculate total points
        total_points = marking_df['Points'].sum()
    
        # Create filename
        exam_title = config.get("exam", {}).get("title", "Exam").replace(" ", "_")
        if timestamp is None:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
        filename = f"{exam_title}_marking_sheet_{timestamp}.xlsx"
        output_path = self.output_dir / filename
    
        # First create a simple version with pandas
        # Create the marking sheet with the total row already included
        compact_df = marking_df[['Question #', 'Correct Answer Letter', 'Points']].copy()
    
        # Add total row as a DataFrame row
        total_row_df = pd.DataFrame([{
            'Question #': 'Total', 
            'Correct Answer Letter': '', 
            'Points': total_points
        }])
    
        # Combine with the compact dataframe
        compact_with_total = pd.concat([compact_df, total_row_df], ignore_index=True)
    
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write compact sheet with total row
            compact_with_total.to_excel(writer, sheet_name='Marking Sheet', index=False)
            
            # Write detailed sheet
            marking_df.to_excel(writer, sheet_name='Detailed', index=False)
            
            # Create grading dataframe
            grade_ranges = [
                (0.9 * total_points, total_points, "A"),
                (0.8 * total_points, 0.9 * total_points - 0.01, "B"),
                (0.7 * total_points, 0.8 * total_points - 0.01, "C"),
                (0.6 * total_points, 0.7 * total_points - 0.01, "D"),
                (0, 0.6 * total_points - 0.01, "F")
            ]
            
            grade_data = []
            for min_pts, max_pts, grade in grade_ranges:
                grade_data.append({
                    'Points Range': f"{min_pts:.1f} - {max_pts:.1f}",
                    'Grade': grade
                })
                
            grade_df = pd.DataFrame(grade_data)
            grade_df.to_excel(writer, sheet_name='Grading Table', index=False)
        
        # Now format the excel file using openpyxl directly
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    
        # Load the workbook to apply formatting
        wb = load_workbook(output_path)
    
        # Format marking sheet
        ws = wb["Marking Sheet"]
    
        # Format headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
        # Format total row
        total_row_idx = len(compact_df) + 1
        for cell in ws[total_row_idx + 1]:  # +1 because openpyxl is 1-indexed
            cell.font = header_font
            cell.fill = header_fill
    
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
        # Format detailed sheet
        ws_detailed = wb["Detailed"]
    
        # Format headers
        for cell in ws_detailed[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
        # Adjust column widths
        ws_detailed.column_dimensions['A'].width = 12  # Question #
        ws_detailed.column_dimensions['B'].width = 15  # Question Type
        ws_detailed.column_dimensions['C'].width = 20  # Topic
        ws_detailed.column_dimensions['D'].width = 15  # Correct Answer #
        ws_detailed.column_dimensions['E'].width = 20  # Correct Answer Letter
        ws_detailed.column_dimensions['F'].width = 30  # Correct Answer Text
        ws_detailed.column_dimensions['G'].width = 10  # Points
        ws_detailed.column_dimensions['H'].width = 50  # Question Text
    
        # Format grading table
        ws_grade = wb["Grading Table"]
    
        # Format headers
        for cell in ws_grade[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
        # Adjust column widths
        ws_grade.column_dimensions['A'].width = 20
        ws_grade.column_dimensions['B'].width = 10
    
        # Save the workbook
        wb.save(output_path)
    
        # Add information about randomization to the marking sheet
        if randomize_answers and answer_mappings:
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                # Create a mapping sheet showing the randomization for each question
                mapping_data = []
                
                for q_num in sorted(answer_mappings.keys()):
                    q_mapping = answer_mappings[q_num]
                    
                    row_data = {'Question #': q_num}
                    
                    # Add original to randomized mapping
                    for orig_pos, rand_letter in q_mapping.items():
                        row_data[f'Orig {chr(64+orig_pos)}'] = rand_letter
                        
                    mapping_data.append(row_data)
                    
                if mapping_data:
                    mapping_df = pd.DataFrame(mapping_data)
                    mapping_df.to_excel(writer, sheet_name='Answer Randomization', index=False)
    
        logger.info(f"Generated marking sheet: {output_path}" + 
                    (" with randomized answer support" if randomize_answers else ""))
    
        return output_path