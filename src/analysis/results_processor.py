import pandas as pd
import numpy as np
from pathlib import Path
import logging
from typing import Dict, Tuple, Optional, List
import argparse
import re
from src.utils import ensure_directory

logger = logging.getLogger(__name__)

class ExamResultsProcessor:
    """Process exam results and generate summary statistics."""
    
    def __init__(self, output_dir: Path = Path("output")):
        """Initialize the exam results processor.
        
        Args:
            output_dir: Directory to save output files
        """
        self.output_dir = ensure_directory(Path(output_dir))
        logger.info(f"Using output directory: {self.output_dir}")
    
    def load_results(self, results_path: Path) -> pd.DataFrame:
        """Load student exam results from Excel file.
        
        Args:
            results_path: Path to Excel file with student answers
            
        Returns:
            DataFrame with student results
        """
        try:
            if results_path.suffix.lower() == ".csv":
                results_df = self._load_csv_results(results_path)
            else:
                results_df = pd.read_excel(results_path)
            logger.info(f"Loaded results from {results_path}")
            
            # Handle the case where the first column is labeled "Question" and contains row numbers
            if results_df.columns[0] == "Question":
                # Make sure the question column contains integers
                results_df["Question"] = results_df["Question"].astype(int)
                
                # If the values start from 0, we might need to add 1 to match marking sheet
                # Uncomment the following line if question numbers should start from 1 instead of 0
                # results_df["Question"] = results_df["Question"] + 1
                
            return results_df
        except Exception as e:
            logger.error(f"Error loading results file: {e}")
            raise

    def _load_csv_results(self, results_path: Path) -> pd.DataFrame:
        """Load and normalize semicolon-delimited exam exports.

        The grading pipeline expects one row per question and one column per student.
        Some exports instead contain one row per student and question ids in headers
        like ``A0`` or ``RF60``. This method transposes that layout.
        """
        raw_df = pd.read_csv(results_path, sep=";", dtype=str, keep_default_na=False)

        if "StudentNumber" not in raw_df.columns:
            raise ValueError("CSV results file must contain a 'StudentNumber' column")

        question_columns = []
        for column_name in raw_df.columns:
            if column_name in {"StudentNumber", "Version"}:
                continue

            match = re.search(r"(\d+)$", str(column_name))
            if match:
                question_columns.append((column_name, int(match.group(1))))

        if not question_columns:
            raise ValueError("CSV results file does not contain any question columns")

        # Some exports include a first row with the answer key and a blank student id.
        student_rows = raw_df[raw_df["StudentNumber"].astype(str).str.strip() != ""].copy()
        if student_rows.empty:
            raise ValueError("CSV results file does not contain any student rows")

        normalized = {"Question": [question_num for _, question_num in question_columns]}

        for _, student_row in student_rows.iterrows():
            student_id = str(student_row["StudentNumber"]).strip()
            normalized[student_id] = [
                str(student_row[column_name]).strip().upper() or np.nan
                for column_name, _ in question_columns
            ]

        return pd.DataFrame(normalized)
    
    def load_marking_sheet(self, marking_sheet_path: Path) -> pd.DataFrame:
        """Load marking sheet with correct answers.
        
        Args:
            marking_sheet_path: Path to marking sheet Excel file
            
        Returns:
            DataFrame with correct answers and points
        """
        try:
            # Try to load the marking sheet
            try:
                # First try with the "Marking Sheet" sheet name
                marking_df = pd.read_excel(marking_sheet_path, sheet_name="Marking Sheet")
            except Exception:
                # If that fails, try the default sheet (first sheet)
                marking_df = pd.read_excel(marking_sheet_path)
                logger.info("Used default sheet instead of 'Marking Sheet'")
            
            # Check for expected columns
            expected_columns = ["Question #", "Correct Answer Letter", "Points"]
            for col in expected_columns:
                if col not in marking_df.columns:
                    # Try to find similar columns
                    if "Question" in marking_df.columns and col == "Question #":
                        marking_df = marking_df.rename(columns={"Question": "Question #"})
                    elif col == "Correct Answer Letter":
                        candidate_columns = [
                            column_name for column_name in marking_df.columns
                            if "answer" in str(column_name).lower() or "key" in str(column_name).lower()
                        ]
                        if candidate_columns:
                            marking_df = marking_df.rename(
                                columns={candidate_columns[0]: "Correct Answer Letter"}
                            )
                    else:
                        raise ValueError(f"Required column '{col}' not found in marking sheet")
            
            # Filter out the total row
            if "Total" in marking_df["Question #"].astype(str).values:
                marking_df = marking_df[marking_df["Question #"].astype(str) != "Total"]
            
            # Convert question numbers to integers if they're not already
            if marking_df["Question #"].dtype != int:
                # Handle cases where question numbers might be strings with "Question" prefix
                try:
                    marking_df["Question #"] = marking_df["Question #"].astype(int)
                except ValueError:
                    # Try to extract numbers from strings like "Question 1"
                    marking_df["Question #"] = marking_df["Question #"].apply(
                        lambda x: int(re.search(r'\d+', str(x)).group()) if re.search(r'\d+', str(x)) else x
                    )
                    marking_df["Question #"] = marking_df["Question #"].astype(int)
            
            # If the first question is 1, but our results file starts from 0, adjust
            # This is a heuristic - may need to be adjusted based on specific formats
            if marking_df["Question #"].min() == 1 and marking_df["Question #"].max() == 19:
                logger.info("Adjusting question numbers to start from 0 instead of 1")
                marking_df["Question #"] = marking_df["Question #"] - 1
                    
            logger.info(f"Loaded marking sheet from {marking_sheet_path}")
            return marking_df
            
        except Exception as e:
            logger.error(f"Error loading marking sheet: {e}")
            raise
    
    def calculate_student_scores(self, 
                            results_df: pd.DataFrame, 
                            marking_df: pd.DataFrame) -> pd.DataFrame:
        """Calculate scores for each student based on marking sheet.
        
        Args:
            results_df: DataFrame with student answers
            marking_df: DataFrame with correct answers and points
            
        Returns:
            DataFrame with student scores
        """
        # Create a new dataframe for student scores
        student_scores = pd.DataFrame()
        
        # Get the question number column
        question_num_col = results_df.columns[0]  # This will be "Question" from the example
        
        # Create a dictionary mapping question numbers to correct answers and points
        marking_dict = marking_df.set_index("Question #").to_dict()
        correct_answers = marking_dict.get("Correct Answer Letter", {})
        question_points = marking_dict.get("Points", {})
        
        logger.info("Question numbers in results: %s", self._format_question_numbers(results_df[question_num_col]))
        logger.info("Question numbers in marking sheet: %s", self._format_question_numbers(marking_df["Question #"]))

        graded_results_df = self._filter_to_marked_questions(
            results_df,
            question_num_col,
            set(correct_answers.keys()),
        )
        
        # Process each student column
        for student_col in results_df.columns[1:]:
            # Initialize points for this student
            points = []
            
            # Process each question
            for _, row in graded_results_df.iterrows():
                question_num = row[question_num_col]
                student_answer = row[student_col]
                
                # Convert student answer to uppercase string for comparison
                if pd.notna(student_answer):
                    student_answer = str(student_answer).strip().upper()
                else:
                    student_answer = ""
                    
                # Check if the answer is correct
                correct_answer = correct_answers.get(question_num, "")
                if isinstance(correct_answer, str):
                    correct_answer = correct_answer.strip().upper()
                    
                question_point_value = question_points.get(question_num, 1)
                
                # Award points if correct
                if student_answer and student_answer == correct_answer:
                    points.append(question_point_value)
                else:
                    points.append(0)
                
                # Debug for troubleshooting
                if question_num in [0, 1]:  # Check specific questions
                    logger.debug(f"Q{question_num} - Student: {student_answer}, Correct: {correct_answer}, Points: {points[-1]}")
            
            # Add student's points to the results
            student_scores[f"{student_col}_points"] = points
            # Calculate total score for this student
            student_scores[f"{student_col}_total"] = sum(points)
        
        # Add question numbers and correct answers for reference
        student_scores["Question #"] = graded_results_df[question_num_col]
        student_scores["Correct Answer"] = student_scores["Question #"].map(correct_answers)
        student_scores["Points Possible"] = student_scores["Question #"].map(question_points)
        
        return student_scores
        
    def generate_question_statistics(self, 
                                results_df: pd.DataFrame, 
                                marking_df: pd.DataFrame, 
                                student_scores: pd.DataFrame) -> pd.DataFrame:
        """Generate statistics about each question's performance.
        
        Args:
            results_df: DataFrame with student answers
            marking_df: DataFrame with correct answers and points
            student_scores: DataFrame with calculated student scores
            
        Returns:
            DataFrame with question statistics
        """
        # Get the question number column
        question_num_col = results_df.columns[0]
        
        # Get number of students
        num_students = len(results_df.columns) - 1  # Subtract 1 for the question number column
        
        # Create a dictionary mapping question numbers to correct answers
        marking_dict = marking_df.set_index("Question #").to_dict()
        correct_answers = marking_dict.get("Correct Answer Letter", {})
        question_points = marking_dict.get("Points", {})
        graded_results_df = self._filter_to_marked_questions(
            results_df,
            question_num_col,
            set(correct_answers.keys()),
            log_ignored=False,
        )
        
        # Initialize statistics
        stats = []
        
        # Process each question
        for _, row in graded_results_df.iterrows():
            question_num = row[question_num_col]
            correct_answer = correct_answers.get(question_num, "")
            point_value = question_points.get(question_num, 1)
            
            # Count responses for each possible answer
            response_counts = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'blank': 0, 'other': 0}
            
            for student_col in results_df.columns[1:]:
                student_answer = row[student_col]
                if pd.isna(student_answer) or str(student_answer).strip() == "":
                    response_counts['blank'] += 1
                else:
                    answer_key = str(student_answer).strip().upper()
                    if answer_key in response_counts:
                        response_counts[answer_key] += 1
                    else:
                        # Handle invalid responses
                        response_counts['other'] += 1
                        logger.warning(f"Invalid response '{student_answer}' for question {question_num}")
            
            # Calculate number of correct responses
            num_correct = response_counts.get(correct_answer, 0)
            correct_pct = (num_correct / num_students) * 100 if num_students > 0 else 0
            
            # Calculate discrimination index (point-biserial correlation)
            # This measures how well a question discriminates between high and low performers
            discrimination = self._calculate_discrimination(question_num, student_scores)
            
            # Compile statistics for this question
            stats.append({
                "Question #": question_num,
                "Correct Answer": correct_answer,
                "Points": point_value, 
                "# Correct": num_correct,
                "% Correct": correct_pct,
                "A": response_counts.get('A', 0),
                "B": response_counts.get('B', 0),
                "C": response_counts.get('C', 0),
                "D": response_counts.get('D', 0),
                "Blank": response_counts.get('blank', 0),
                "Other": response_counts.get('other', 0),
                "Discrimination": discrimination
            })
        
        # Create DataFrame from statistics
        stats_df = pd.DataFrame(stats)
        
        return stats_df

    def _filter_to_marked_questions(
        self,
        results_df: pd.DataFrame,
        question_num_col: str,
        marked_questions: set,
        log_ignored: bool = True,
    ) -> pd.DataFrame:
        """Keep only result rows that have a corresponding marking-sheet answer."""
        mask = results_df[question_num_col].isin(marked_questions)
        ignored_questions = sorted(set(results_df.loc[~mask, question_num_col]))
        if ignored_questions and log_ignored:
            logger.info(
                "Ignoring %s unmarked result columns/questions: %s",
                len(ignored_questions),
                ignored_questions,
            )
        return results_df.loc[mask].copy()

    @staticmethod
    def _format_question_numbers(question_numbers: pd.Series) -> List[int]:
        """Return sorted question numbers as native Python integers for readable logs."""
        return sorted(int(question_number) for question_number in question_numbers.unique())
    
    def _calculate_discrimination(self, question_num: int, student_scores: pd.DataFrame) -> float:
        """Calculate discrimination index for a question.
        
        Args:
            question_num: Question number
            student_scores: DataFrame with student scores
            
        Returns:
            Discrimination index (-1 to 1, higher means better discrimination)
        """
        # Get rows for this question
        question_data = student_scores[student_scores["Question #"] == question_num]
        
        if len(question_data) == 0:
            return 0
        
        # Get student columns (those containing _points but not _total)
        point_cols = [col for col in student_scores.columns if "_points" in col and "_total" not in col]
        total_cols = [col for col in student_scores.columns if "_total" in col]
        
        if not point_cols or not total_cols:
            return 0
            
        # Get point values for this question for each student
        point_values = []
        total_scores = []
        
        for i, point_col in enumerate(point_cols):
            if i < len(total_cols):
                # Get corresponding total score column
                total_col = total_cols[i]
                
                # Get this question's points for the student
                if not question_data.empty and point_col in question_data.columns:
                    points = question_data[point_col].iloc[0]
                    
                    # Get student's total score excluding this question
                    if not student_scores.empty and total_col in student_scores.columns:
                        total = student_scores[total_col].max()
                        adjusted_total = total - points
                        
                        point_values.append(points)
                        total_scores.append(adjusted_total)
        
        # Calculate correlation coefficient if possible
        if len(point_values) > 1 and len(total_scores) > 1:
            try:
                # Check for zero standard deviation which causes division by zero
                if np.std(point_values) > 0 and np.std(total_scores) > 0:
                    return np.corrcoef(point_values, total_scores)[0, 1]
                else:
                    return 0
            except:
                return 0
        return 0
    
    def generate_summary_report(self,
                            student_scores: pd.DataFrame,
                            question_stats: pd.DataFrame,
                            output_name: str,
                            results_df: Optional[pd.DataFrame] = None) -> Path:
        """Generate and save summary report with student scores and question statistics.
        
        Args:
            student_scores: DataFrame with student scores
            question_stats: DataFrame with question statistics
            output_name: Base name for output file
            results_df: Optional DataFrame with normalized student answers
            
        Returns:
            Path to saved report
        """
        student_summary = self._build_student_summary(student_scores)
        question_report = self._build_question_report(question_stats)
        overview = self._build_overview(student_summary, question_report, student_scores)
        score_matrix = self._build_score_matrix(student_scores)
        answer_matrix = self._build_answer_matrix(results_df, student_scores)

        output_path = self.output_dir / f"{output_name}_grading_report.xlsx"

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview.to_excel(writer, sheet_name="Overview", index=False)
            student_summary.to_excel(writer, sheet_name="Student Scores", index=False)
            question_report.to_excel(writer, sheet_name="Question Analysis", index=False)
            answer_matrix.to_excel(writer, sheet_name="Answer Matrix", index=False)
            score_matrix.to_excel(writer, sheet_name="Score Matrix", index=False)

            self._format_report_workbook(writer.book, student_summary, question_report)

        logger.info(f"Generated summary report: {output_path}")
        
        return output_path

    def _build_student_summary(self, student_scores: pd.DataFrame) -> pd.DataFrame:
        """Build one row per student with ranking and score percentages."""
        total_cols = [col for col in student_scores.columns if col.endswith("_total")]
        total_possible = student_scores["Points Possible"].sum() if "Points Possible" in student_scores.columns else np.nan

        rows = []
        for col in total_cols:
            student_id = col.replace("_total", "")
            total_score = student_scores[col].max()
            percentage = (total_score / total_possible) * 100 if total_possible else np.nan
            rows.append({
                "Rank": np.nan,
                "Student": student_id,
                "Total Points": total_score,
                "Percentage": percentage,
            })

        summary = pd.DataFrame(rows)
        if summary.empty:
            return pd.DataFrame(columns=["Rank", "Student", "Total Points", "Percentage", "Percentile"])

        summary = summary.sort_values(["Total Points", "Student"], ascending=[False, True]).reset_index(drop=True)
        summary["Rank"] = summary["Total Points"].rank(method="min", ascending=False).astype(int)
        if len(summary) > 1:
            summary["Percentile"] = ((len(summary) - summary["Rank"]) / (len(summary) - 1)) * 100
        else:
            summary["Percentile"] = 100.0

        return summary[["Rank", "Student", "Total Points", "Percentage", "Percentile"]]

    def _build_question_report(self, question_stats: pd.DataFrame) -> pd.DataFrame:
        """Add interpretation columns to the question-level statistics."""
        report = question_stats.copy()
        if report.empty:
            return report

        report = report.sort_values("Question #").reset_index(drop=True)
        if "Other" not in report.columns:
            report["Other"] = 0

        total_responses = report[["A", "B", "C", "D", "Blank", "Other"]].sum(axis=1)
        report["Answered"] = total_responses - report["Blank"]
        report["Response Rate"] = np.where(total_responses > 0, (report["Answered"] / total_responses) * 100, 0)
        report["Difficulty"] = report["% Correct"].apply(self._difficulty_label)
        report["Most Common Wrong"] = report.apply(self._most_common_wrong_answer, axis=1)
        return report[
            [
                "Question #", "Correct Answer", "Points", "# Correct", "% Correct",
                "Difficulty", "Discrimination", "Response Rate", "Most Common Wrong",
                "A", "B", "C", "D", "Blank", "Other",
            ]
        ]

    def _build_overview(
        self,
        student_summary: pd.DataFrame,
        question_report: pd.DataFrame,
        student_scores: pd.DataFrame,
    ) -> pd.DataFrame:
        """Create a compact metrics sheet for the front of the workbook."""
        total_possible = student_scores["Points Possible"].sum() if "Points Possible" in student_scores.columns else np.nan
        metrics = [
            ("Students", len(student_summary)),
            ("Questions", len(question_report)),
            ("Total possible points", total_possible),
            ("Mean points", student_summary["Total Points"].mean()),
            ("Median points", student_summary["Total Points"].median()),
            ("Standard deviation", student_summary["Total Points"].std()),
            ("Minimum points", student_summary["Total Points"].min()),
            ("Maximum points", student_summary["Total Points"].max()),
            ("Mean percentage", student_summary["Percentage"].mean()),
            ("Median percentage", student_summary["Percentage"].median()),
            ("Questions below 50% correct", int((question_report["% Correct"] < 50).sum())),
            ("Questions above 90% correct", int((question_report["% Correct"] >= 90).sum())),
        ]
        return pd.DataFrame(metrics, columns=["Metric", "Value"])

    def _format_report_workbook(
        self,
        workbook,
        student_summary: pd.DataFrame,
        question_report: pd.DataFrame,
    ) -> None:
        """Apply presentation formatting and embedded workbook charts."""
        from openpyxl.chart import BarChart, LineChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        accent_fill = PatternFill("solid", fgColor="D9EAF7")

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                width = min(max(max_length + 2, 12), 32)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "0.00"
                    if worksheet.cell(row=1, column=cell.column).value in {"Percentage", "Percentile", "% Correct", "Response Rate"}:
                        cell.number_format = "0.0"

        overview_sheet = workbook["Overview"]
        overview_sheet.freeze_panes = None
        overview_sheet["A1"].fill = accent_fill
        overview_sheet["B1"].fill = accent_fill

        if not student_summary.empty:
            student_sheet = workbook["Student Scores"]
            score_chart = BarChart()
            score_chart.title = "Student Score Percentages"
            score_chart.y_axis.title = "Percent"
            score_chart.x_axis.title = "Student"
            score_chart.height = 9
            score_chart.width = 18
            data = Reference(student_sheet, min_col=4, min_row=1, max_row=len(student_summary) + 1)
            categories = Reference(student_sheet, min_col=2, min_row=2, max_row=len(student_summary) + 1)
            score_chart.add_data(data, titles_from_data=True)
            score_chart.set_categories(categories)
            overview_sheet.add_chart(score_chart, "D2")

        if not question_report.empty:
            question_sheet = workbook["Question Analysis"]
            question_chart = LineChart()
            question_chart.title = "Question Correctness"
            question_chart.y_axis.title = "Percent Correct"
            question_chart.x_axis.title = "Question"
            question_chart.height = 9
            question_chart.width = 18
            data = Reference(question_sheet, min_col=5, min_row=1, max_row=len(question_report) + 1)
            categories = Reference(question_sheet, min_col=1, min_row=2, max_row=len(question_report) + 1)
            question_chart.add_data(data, titles_from_data=True)
            question_chart.set_categories(categories)
            overview_sheet.add_chart(question_chart, "D20")

    def _build_score_matrix(self, student_scores: pd.DataFrame) -> pd.DataFrame:
        """Return points per question with one column per student."""
        matrix = pd.DataFrame({
            "Question #": student_scores["Question #"],
            "Correct Answer": student_scores["Correct Answer"],
            "Points Possible": student_scores["Points Possible"],
        })

        for col in [col for col in student_scores.columns if col.endswith("_points")]:
            matrix[col.replace("_points", "")] = student_scores[col]

        return matrix

    def _build_answer_matrix(
        self,
        results_df: Optional[pd.DataFrame],
        student_scores: pd.DataFrame,
    ) -> pd.DataFrame:
        """Return normalized answers with correct answers and point values."""
        if results_df is None:
            return self._build_score_matrix(student_scores)

        question_col = results_df.columns[0]
        graded_questions = set(student_scores["Question #"])
        results_df = results_df[results_df[question_col].isin(graded_questions)].reset_index(drop=True)
        student_scores = student_scores.reset_index(drop=True)
        matrix = pd.DataFrame({
            "Question #": results_df[question_col],
            "Correct Answer": student_scores["Correct Answer"],
            "Points Possible": student_scores["Points Possible"],
        })

        for col in results_df.columns[1:]:
            matrix[col] = results_df[col]

        return matrix

    @staticmethod
    def _difficulty_label(correct_pct: float) -> str:
        if correct_pct < 50:
            return "Hard"
        if correct_pct >= 85:
            return "Easy"
        return "Moderate"

    @staticmethod
    def _most_common_wrong_answer(row: pd.Series) -> str:
        correct_answer = str(row["Correct Answer"]).strip().upper()
        counts = {
            answer: row.get(answer, 0)
            for answer in ["A", "B", "C", "D"]
            if answer != correct_answer
        }
        if not counts:
            return ""
        answer, count = max(counts.items(), key=lambda item: item[1])
        return f"{answer} ({count})" if count else ""
    
def main():
    """Main entry point for results processor."""
    # Configure argument parser
    parser = argparse.ArgumentParser(description="Process exam results")
    parser.add_argument("results_file", help="Path to Excel file with student answers")
    parser.add_argument("marking_sheet", help="Path to marking sheet Excel file")
    parser.add_argument("--output-dir", default="output", help="Directory to save output files")
    parser.add_argument("--output-name", default="exam_results", help="Base name for output files")
    
    # Parse arguments
    args = parser.parse_args()
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    
    # Process results
    processor = ExamResultsProcessor(output_dir=Path(args.output_dir))
    
    # Load data
    results_df = processor.load_results(Path(args.results_file))
    marking_df = processor.load_marking_sheet(Path(args.marking_sheet))
    
    # Calculate scores
    student_scores = processor.calculate_student_scores(results_df, marking_df)
    
    # Generate question statistics
    question_stats = processor.generate_question_statistics(
        results_df, marking_df, student_scores
    )
    
    # Generate and save summary report
    output_path = processor.generate_summary_report(
        student_scores, question_stats, args.output_name, results_df=results_df
    )
    
    print(f"Results processed successfully. Summary report saved to: {output_path}")

if __name__ == "__main__":
    main()
